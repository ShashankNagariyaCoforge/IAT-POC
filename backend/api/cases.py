"""
Cases API endpoints (Step 15).
Provides all read-only endpoints for the React UI.
All routes require JWT authentication (enforced by middleware).

DEMO MODE: When DEMO_MODE=true in .env, uses LocalDBService (TinyDB) instead
           of CosmosDBService, and reads extracted text from local files.
"""

import logging
import os
from typing import Optional, Union

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import Response, StreamingResponse
import fitz
import io
from pydantic import BaseModel
from typing import List, Any, Dict

from config import settings

logger = logging.getLogger(__name__)
router = APIRouter()


def _get_cosmos() -> Union["CosmosDBService", "LocalDBService"]:  # type: ignore[name-defined]
    """Dependency: returns DB service — LocalDB in demo mode, Cosmos otherwise."""
    if settings.demo_mode:
        from services.local_db import LocalDBService
        return LocalDBService()
    from services.cosmos_db import CosmosDBService
    return CosmosDBService()


@router.get("/cases")
async def list_cases(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=200),
    search: Optional[str] = Query(default=None, description="Search by Case ID, sender, or subject"),
    category: Optional[str] = Query(default=None),
    status: Optional[str] = Query(default=None),
    requires_human_review: Optional[bool] = Query(default=None),
    date_from: Optional[str] = Query(default=None, description="ISO date string"),
    date_to: Optional[str] = Query(default=None, description="ISO date string"),
    sort_by: str = Query(default="created_at"),
    sort_order: str = Query(default="DESC", pattern="^(ASC|DESC)$"),
):
    """
    List cases with filtering, sorting, and pagination.

    Returns paginated case list for the UI home screen.
    """
    cosmos = _get_cosmos()
    result = await cosmos.list_cases(
        page=page,
        page_size=page_size,
        search=search,
        category=category,
        status=status,
        requires_human_review=requires_human_review,
        date_from=date_from,
        date_to=date_to,
        sort_by=sort_by,
        sort_order=sort_order,
    )
    return result

@router.get("/cases/dashboard-metrics")
async def get_dashboard_metrics():
    """Returns aggregated metrics for the dashboard view."""
    if settings.demo_mode:
        from services.local_db import LocalDBService
        db = LocalDBService()
    else:
        from services.cosmos_db import CosmosDBService
        db = CosmosDBService()
        
    metrics = await db.get_dashboard_metrics()
    return metrics


@router.get("/cases/{case_id}")
async def get_case(case_id: str):
    """
    Get full case detail by Case ID.
    Used for the Case Detail page.
    """
    cosmos = _get_cosmos()
    case = await cosmos.get_case(case_id)
    if not case:
        raise HTTPException(status_code=404, detail=f"Case not found: {case_id}")
    return case


@router.get("/cases/{case_id}/emails")
async def get_case_emails(case_id: str):
    """
    Get all emails in a case chain, ordered chronologically.
    Note: email bodies are PII-masked before being stored.
    """
    cosmos = _get_cosmos()
    case = await cosmos.get_case(case_id)
    if not case:
        raise HTTPException(status_code=404, detail=f"Case not found: {case_id}")
    emails = await cosmos.get_emails_for_case(case_id)
    return {"emails": emails, "total": len(emails)}


@router.get("/cases/{case_id}/documents")
async def get_case_documents(case_id: str):
    """
    Get all documents associated with a case.
    Includes extracted text previews (PII masked) and processing metadata.
    """
    cosmos = _get_cosmos()
    case = await cosmos.get_case(case_id)
    if not case:
        raise HTTPException(status_code=404, detail=f"Case not found: {case_id}")

    docs = await cosmos.get_documents_for_case(case_id)

    # Fetch text previews
    enriched = []
    for doc in docs:
        preview = None
        if settings.demo_mode:
            # Read from local extracted_text folder
            local_path = doc.get("extracted_text_local_path")
            if local_path and os.path.exists(local_path):
                try:
                    with open(local_path, "r", encoding="utf-8") as fp:
                        preview = fp.read(500)
                except Exception:
                    pass
        else:
            # Read from Azure Blob
            from services.blob_storage import BlobStorageService
            blob = BlobStorageService()
            text_path = doc.get("extracted_text_blob_path")
            if text_path:
                try:
                    container, blob_name = text_path.split("/", 1)
                    text = await blob.download_text(container, blob_name)
                    preview = text[:500]
                except Exception:
                    pass
        
        # Ensure 'file_name' exists for frontend compatibility
        doc_data = {**doc, "extracted_text_preview": preview}
        if "filename" in doc_data and "file_name" not in doc_data:
            doc_data["file_name"] = doc_data["filename"]
            
        enriched.append(doc_data)

    return {"documents": enriched, "total": len(enriched)}


@router.get("/cases/{case_id}/documents/{document_id}/pdf")
async def get_case_document_pdf(case_id: str, document_id: str):
    """
    Get the raw PDF bytes for a specific document.
    """
    cosmos = _get_cosmos()
    case = await cosmos.get_case(case_id)
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")

    docs = await cosmos.get_documents_for_case(case_id)
    doc = next((d for d in docs if d.get("document_id") == document_id), None)
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    blob_path = doc.get("blob_path")
    
    if not blob_path and settings.demo_mode:
        # In demo mode, documents may not have a blob_path. They might just be stored locally.
        # Fallback to checking local path if we have it
        local_path = doc.get("local_path") or doc.get("extracted_text_local_path")
        if local_path and os.path.exists(local_path):
             # If it's a txt file, this is just a fallback. Usually we want the real PDF.
             pass

    if not blob_path:
        raise HTTPException(status_code=404, detail="PDF blob path not found in document record")

    try:
        from services.blob_storage import BlobStorageService
        blob = BlobStorageService()
        
        # In this architecture, blob_path is relative to the container.
        # It does NOT include the container name.
        container = settings.blob_container_raw_emails
        blob_name = blob_path
        
        pdf_bytes = await blob.download_bytes(container, blob_name)
        return Response(content=pdf_bytes, media_type="application/pdf")
    except Exception as e:
        logger.error(f"Failed to download PDF {blob_path}: {e}")
        raise HTTPException(status_code=500, detail="Failed to download PDF from storage")


@router.get("/cases/{case_id}/documents/{document_id}/annotated")
async def get_case_document_annotated_pdf(case_id: str, document_id: str):
    """
    Get the pre-rendered annotated PDF for a specific document.
    """
    cosmos = _get_cosmos()
    classification = await cosmos.get_classification_for_case(case_id)
    if not classification:
        raise HTTPException(status_code=404, detail="Classification not found")
        
    annotated_docs = classification.get("annotated_docs", {})
    blob_path = annotated_docs.get(document_id)
    
    if not blob_path:
        # Fallback to original PDF if annotated version doesn't exist yet
        return await get_case_document_pdf(case_id, document_id)

    try:
        from services.blob_storage import BlobStorageService
        blob = BlobStorageService()
        container = settings.blob_container_raw_emails  # Must match upload container in process.py
        print(f"DEBUG [cases/annotated]: Downloading from container='{container}', blob='{blob_path}'")
        pdf_bytes = await blob.download_bytes(container, blob_path)
        print(f"DEBUG [cases/annotated]: Downloaded {len(pdf_bytes)} bytes successfully")
        return Response(content=pdf_bytes, media_type="application/pdf")
    except Exception as e:
        logger.error(f"Failed to download annotated PDF {blob_path}: {e}")
        # Fallback to original PDF if annotated version can't be retrieved
        return await get_case_document_pdf(case_id, document_id)


@router.get("/cases/{case_id}/documents/{document_id}/view")
async def get_case_document_view(case_id: str, document_id: str):
    """
    Generic document viewer endpoint.
    - PDF / images → served with correct content-type (for iframe / react-pdf)
    - XLSX / XLS   → converted to a styled HTML table for inline display
    - DOCX         → served as-is (browser will download; future: convert to PDF)
    """
    cosmos = _get_cosmos()
    docs = await cosmos.get_documents_for_case(case_id)
    doc = next((d for d in docs if d.get("document_id") == document_id), None)
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    blob_path = doc.get("blob_path")
    if not blob_path:
        raise HTTPException(status_code=404, detail="Blob path not found")

    filename = (doc.get("file_name") or doc.get("filename") or blob_path or "").lower()
    ext = filename.rsplit(".", 1)[-1] if "." in filename else ""

    from services.blob_storage import BlobStorageService
    blob = BlobStorageService()
    container = settings.blob_container_raw_emails

    try:
        raw_bytes = await blob.download_bytes(container, blob_path)
    except Exception as e:
        logger.error(f"Failed to download document {blob_path}: {e}")
        raise HTTPException(status_code=500, detail="Failed to download document")

    # ── XLSX / XLS → HTML table ──────────────────────────────────────────────
    if ext in ("xlsx", "xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
            tabs_html = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                header = rows[0]
                data_rows = rows[1:]
                th_cells = "".join(
                    f"<th>{(str(c) if c is not None else '')}</th>" for c in header
                )
                tr_rows = "".join(
                    "<tr>" + "".join(
                        f"<td>{(str(c) if c is not None else '')}</td>" for c in row
                    ) + "</tr>"
                    for row in data_rows
                )
                tabs_html.append(
                    f"<h3 class='sheet-name'>{sheet_name}</h3>"
                    f"<div class='tbl-wrap'><table><thead><tr>{th_cells}</tr></thead>"
                    f"<tbody>{tr_rows}</tbody></table></div>"
                )

            html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8"/>
<style>
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
          font-size: 12px; margin: 0; padding: 12px; background: #f8fafc; color: #1e293b; }}
  .sheet-name {{ font-size: 13px; font-weight: 700; color: #475569; margin: 16px 0 6px; }}
  .tbl-wrap {{ overflow-x: auto; border-radius: 8px; border: 1px solid #e2e8f0;
               box-shadow: 0 1px 4px rgba(0,0,0,.06); margin-bottom: 24px; }}
  table {{ border-collapse: collapse; min-width: 100%; background: #fff; }}
  thead tr {{ background: #f1f5f9; }}
  th {{ padding: 7px 12px; text-align: left; font-weight: 700; font-size: 11px;
        color: #64748b; border-bottom: 2px solid #e2e8f0; white-space: nowrap; }}
  td {{ padding: 6px 12px; border-bottom: 1px solid #f1f5f9; white-space: nowrap; }}
  tr:last-child td {{ border-bottom: none; }}
  tr:hover td {{ background: #f8fafc; }}
</style>
</head>
<body>{''.join(tabs_html) if tabs_html else '<p style="color:#94a3b8">No data found in spreadsheet.</p>'}</body>
</html>"""
            return Response(content=html, media_type="text/html")
        except Exception as e:
            logger.error(f"Failed to convert xlsx {blob_path}: {e}")
            raise HTTPException(status_code=500, detail=f"Failed to render spreadsheet: {e}")

    # ── Images → serve directly ──────────────────────────────────────────────
    image_types = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
                   "tiff": "image/tiff", "tif": "image/tiff", "bmp": "image/bmp",
                   "gif": "image/gif", "webp": "image/webp"}
    if ext in image_types:
        return Response(content=raw_bytes, media_type=image_types[ext])

    # ── DOCX → serve for download (browser handles it) ──────────────────────
    if ext == "docx":
        return Response(
            content=raw_bytes,
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            headers={"Content-Disposition": f"inline; filename=\"{filename}\""},
        )

    # ── Default → PDF ────────────────────────────────────────────────────────
    return Response(content=raw_bytes, media_type="application/pdf")


@router.get("/cases/{case_id}/documents/{document_id}/pages/{page_number}/image")
async def get_case_document_page_image(case_id: str, document_id: str, page_number: int):
    """
    Renders a specific page of a PDF document as a PNG image.
    This is used for high-fidelity SVG annotation overlays.
    """
    try:
        # Re-use existing logic to get PDF bytes
        resp = await get_case_document_pdf(case_id, document_id)
        if not isinstance(resp, Response):
             raise HTTPException(status_code=500, detail="Could not retrieve PDF bytes")
        
        pdf_bytes = resp.body
        
        # Render page using PyMuPDF (fitz)
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        if page_number < 1 or page_number > len(doc):
            raise HTTPException(status_code=400, detail=f"Invalid page number {page_number}. Total pages: {len(doc)}")
            
        page = doc[page_number - 1]
        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2)) # 2x scale for quality
        img_bytes = pix.tobytes("png")
        doc.close()
        
        return Response(content=img_bytes, media_type="image/png")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to render page image: {e}")
        raise HTTPException(status_code=500, detail="Error rendering document page")


@router.get("/cases/{case_id}/classification")
async def get_case_classification(case_id: str):
    """
    Get the classification result for a case.
    Includes confidence score, category, summary, and notification status.
    """
    cosmos = _get_cosmos()
    case = await cosmos.get_case(case_id)
    if not case:
        raise HTTPException(status_code=404, detail=f"Case not found: {case_id}")
    result = await cosmos.get_classification_for_case(case_id)
    if not result:
        return {"classification": None, "message": "Classification not yet available."}
    return {"classification": result}

@router.get("/cases/{case_id}/timeline")
async def get_case_timeline(case_id: str):
    """
    Get the processing event timeline for a case.
    Events: email received, processed, classified, notified.
    """
    cosmos = _get_cosmos()
    case = await cosmos.get_case(case_id)
    if not case:
        raise HTTPException(status_code=404, detail=f"Case not found: {case_id}")
    timeline = await cosmos.get_timeline_for_case(case_id)
    return {"timeline": timeline}


@router.get("/cases/{case_id}/pipeline-status")
async def get_case_pipeline_status(case_id: str):
    """
    Returns agent pipeline status derived from the case's current status.

    Agents:
      0: Orchestrator Agent   — always runs first
      1: Email Agent          — EmailFetcherService + email.json download
      2: PII Agent            — PIIMasker
      3: Content Safety Agent — ContentSafetyService
      4: Classification Agent — Classifier
      5: Extraction Agent     — DI spatial mapping
      6: Enrichment Agent     — Web crawling + AI enrichment (runs parallel with 2-5)
    """
    cosmos = _get_cosmos()
    case = await cosmos.get_case(case_id)
    if not case:
        raise HTTPException(status_code=404, detail=f"Case not found: {case_id}")

    status = case.get("status", "RECEIVED")
    pii_skipped = case.get("pii_skipped", False)

    # Base agent definitions
    agents = [
        {"id": "orchestrator", "name": "Orchestrator",      "type": "Master",    "detail": "Coordinating pipeline", "score": 100},
        {"id": "email",        "name": "Email Agent",       "type": "Retrieval", "detail": "Fetching & parsing email", "score": 98},
        {"id": "pii",          "name": "PII Agent",         "type": "Masking",   "detail": "Masking sensitive data", "score": 95},
        {"id": "safety",       "name": "Content Safety",    "type": "Safety",    "detail": "Running safety checks", "score": 97},
        {"id": "classifier",   "name": "Classification",    "type": "Inference", "detail": "AI classification", "score": 0},
        {"id": "extraction",   "name": "Extraction Agent",  "type": "Vision",    "detail": "High-fidelity mapping", "score": 0},
        {"id": "enrichment",   "name": "Enrichment Agent",  "type": "Crawl+AI",  "detail": "Web enrichment & search", "score": 0},
    ]

    # Filter out PII agent if skipped
    if pii_skipped:
        agents = [a for a in agents if a["id"] != "pii"]

    # Determine agent states based on case status
    TERMINAL_STATUSES = {"PROCESSED", "CLASSIFIED", "BLOCKED_SAFETY", "NEEDS_REVIEW_SAFETY", "FAILED", "PENDING_REVIEW"}

    if status == "RECEIVED":
        states = ["active"] + ["pending"] * (len(agents) - 1)
        current_agent_index = 0

    elif status == "PROCESSING":
        # Use pipeline_step for granular real-time progress
        step = case.get("pipeline_step")
        
        # Map step names to their respective agent indices
        if not pii_skipped:
            step_map = {
                "fetch_content": 1, "pii_masking": 2, "content_safety": 3,
                "classification": 4, "extraction": 5, "enrichment": 6, "completed": 6
            }
        else:
            step_map = {
                "fetch_content": 1, "content_safety": 2, "classification": 3,
                "extraction": 4, "enrichment": 5, "completed": 5
            }
            
        current_agent_index = step_map.get(step, 1)
        
        states = []
        for i in range(len(agents)):
            if i < current_agent_index:
                states.append("completed")
            elif i == current_agent_index:
                states.append("active")
            else:
                states.append("pending")
        
        # Enrichment agent (the last one) runs in parallel from PII/Safety onwards
        enrichment_idx = len(agents) - 1
        if 2 <= current_agent_index < enrichment_idx:
            states[enrichment_idx] = "active"

    elif status == "BLOCKED_SAFETY":
        # Orchestrator, Email, PII (if any), Safety(failed)
        safety_idx = 2 if pii_skipped else 3
        states = ["completed"] * safety_idx + ["failed"] + ["pending"] * (len(agents) - safety_idx - 2) + ["completed"]
        current_agent_index = safety_idx
        agents[safety_idx]["detail"] = "Blocked — content policy violation"
        agents[safety_idx]["score"] = 0

    elif status == "NEEDS_REVIEW_SAFETY":
        safety_idx = 2 if pii_skipped else 3
        states = ["completed"] * len(agents)
        states[safety_idx] = "warning"
        current_agent_index = len(agents) - 1 # Enrichment is last
        agents[safety_idx]["detail"] = "Flagged for review"
        
        # Pull indices for classifier/extraction
        classifier_idx = 3 if pii_skipped else 4
        extraction_idx = 4 if pii_skipped else 5
        agents[classifier_idx]["score"] = 85
        agents[extraction_idx]["score"] = 85

    elif status == "FAILED":
        states = ["completed", "failed"] + ["pending"] * (len(agents) - 2)
        current_agent_index = 1
        agents[1]["detail"] = "Processing error"

    elif status in {"PROCESSED", "CLASSIFIED", "PENDING_REVIEW"}:
        states = ["completed"] * len(agents)
        current_agent_index = len(agents) - 1
        
        # Pull indices
        classifier_idx = 3 if pii_skipped else 4
        extraction_idx = 4 if pii_skipped else 5
        enrichment_idx = 5 if pii_skipped else 6

        # Pull confidence score from classification
        classification = await cosmos.get_classification_for_case(case_id)
        if classification:
            score = classification.get("confidence_score", 0)
            agents[classifier_idx]["score"] = int((score or 0) * 100) if (score or 0) <= 1 else int(score or 0)
            agents[classifier_idx]["detail"] = f"Classified: {classification.get('classification_category', 'Unknown')}"
            agents[extraction_idx]["score"] = agents[classifier_idx]["score"]
            agents[extraction_idx]["detail"] = "Coordinates mapped for UI"

        # Pull enrichment score
        enrichment = await cosmos.get_enrichment_for_case(case_id)
        if enrichment:
            enrich_data = enrichment.get("enrichment", {})
            field_keys = [
                "entity_type", "naics_code", "entity_structure",
                "years_in_business", "number_of_employees", "territory_code",
                "limit_of_liability", "deductible",
            ]
            filled = sum(
                1 for k in field_keys
                if enrich_data.get(k) and isinstance(enrich_data[k], dict) and enrich_data[k].get("value")
            )
            total = len(field_keys)
            agents[enrichment_idx]["score"] = int((filled / total) * 100) if total > 0 else 0
            agents[enrichment_idx]["detail"] = f"Enriched {filled}/{total} fields from web"
        else:
            agents[enrichment_idx]["detail"] = "Awaiting enrichment data"

    else:
        states = ["active"] + ["pending"] * (len(agents) - 1)
        current_agent_index = 0

    for i, agent in enumerate(agents):
        if i < len(states):
            agent["status"] = states[i]

    return {
        "case_id": case_id,
        "status": status,
        "agents": agents,
        "current_agent_index": current_agent_index,
        "is_terminal": status in TERMINAL_STATUSES,
    }


# ── GET /cases/{case_id}/enrichment ───────────────────────────────────────────
@router.get("/cases/{case_id}/enrichment")
async def get_case_enrichment(case_id: str):
    """
    Get the enrichment result for a case.
    Returns web-crawled enrichment fields with confidence scores.
    """
    cosmos = _get_cosmos()
    case = await cosmos.get_case(case_id)
    if not case:
        raise HTTPException(status_code=404, detail=f"Case not found: {case_id}")
    result = await cosmos.get_enrichment_for_case(case_id)
    if not result:
        return {"enrichment": None, "message": "Enrichment not yet available."}
    return {"enrichment": result}


# ── PATCH /cases/{case_id}/fields ─────────────────────────────────────────────
class FieldUpdate(BaseModel):
    field_name: str
    value: str

class PatchFieldsRequest(BaseModel):
    fields: List[FieldUpdate]
    updated_by: str = "user"

@router.patch("/cases/{case_id}/fields")
async def patch_case_fields(case_id: str, body: PatchFieldsRequest):
    """
    Save human-edited field overrides for a case.
    Stores in the classification result's extra_fields dict.
    """
    cosmos = _get_cosmos()
    case = await cosmos.get_case(case_id)
    if not case:
        raise HTTPException(status_code=404, detail=f"Case not found: {case_id}")

    classification = await cosmos.get_classification_for_case(case_id)
    if not classification:
        raise HTTPException(status_code=404, detail="Classification not yet available for this case.")

    # Merge new field values into hitl_fields dict
    hitl_fields: Dict[str, Any] = classification.get("hitl_fields", {}) or {}
    for field_update in body.fields:
        hitl_fields[field_update.field_name] = field_update.value

    # Persist back
    await cosmos.update_classification_hitl_fields(case_id, hitl_fields, body.updated_by)

    return {
        "case_id": case_id,
        "updated_fields": [f.field_name for f in body.fields],
        "hitl_fields": hitl_fields,
    }


# ── GET /cases/{case_id}/snapshot ─────────────────────────────────────────────
@router.get("/cases/{case_id}/snapshot")
async def get_case_snapshot(case_id: str):
    """
    Returns a comprehensive snapshot of the case for the read-only archive view.
    Aggregates case, classification, pipeline status, and hitl fields.
    """
    cosmos = _get_cosmos()
    case = await cosmos.get_case(case_id)
    if not case:
        raise HTTPException(status_code=404, detail=f"Case not found: {case_id}")

    classification = await cosmos.get_classification_for_case(case_id)
    pipeline = await get_case_pipeline_status(case_id)

    # Build audit trail from pipeline agents
    audit_steps = []
    for agent in pipeline.get("agents", []):
        conf = agent.get("score", 0)
        audit_steps.append({
            "label": agent["name"].upper(),
            "status": agent["status"],
            "confidence": conf,
            "detail": agent.get("detail", ""),
        })

    # Extracted / HITL fields
    hitl_fields = {}
    extracted_fields = {}
    if classification:
        hitl_fields = classification.get("hitl_fields", {}) or {}
        # Build a flat dict of known classification fields
        extracted_fields = {
            "Document Type":     classification.get("document_type", ""),
            "Category":          classification.get("classification_category", ""),
            "Confidence":        f"{round((classification.get('confidence_score', 0) or 0) * 100)}%",
            "Urgency":           classification.get("urgency", ""),
            "Policy Reference":  classification.get("policy_reference", ""),
            "Claim Type":        classification.get("claim_type", ""),
            "Sender":            case.get("sender", ""),
            "Received":          case.get("created_at", ""),
        }

    return {
        "case_id": case_id,
        "status": case.get("status"),
        "subject": case.get("subject", ""),
        "sender": case.get("sender", ""),
        "created_at": case.get("created_at"),
        "updated_at": case.get("updated_at"),
        "requires_human_review": case.get("requires_human_review", False),
        "classification": classification,
        "pipeline": pipeline,
        "audit_steps": audit_steps,
        "extracted_fields": extracted_fields,
        "hitl_fields": hitl_fields,
    }


@router.get("/cases/{case_id}/download-masked")
async def get_case_masked_report(case_id: str):
    """
    Downloads the PII-masked HTML report for a processed case.
    """
    cosmos = _get_cosmos()
    classification = await cosmos.get_classification_for_case(case_id)
    if not classification or not classification.get("pii_report_blob_path"):
        raise HTTPException(
            status_code=404, 
            detail="Masking report not found. Has this case been processed?"
        )

    blob_path = classification["pii_report_blob_path"]
    try:
        from fastapi import Response
        import os # Added for os.path.exists
        if settings.demo_mode:
            # Read from local file
            if not os.path.exists(blob_path):
                 raise HTTPException(status_code=404, detail="Local report file not found.")
            with open(blob_path, "r", encoding="utf-8") as f:
                content = f.read()
        else:
            # Read from Azure Blob
            from services.blob_storage import BlobStorageService
            blob = BlobStorageService()
            container, name = blob_path.split("/", 1)
            content = await blob.download_text(container, name)

        return Response(
            content=content,
            media_type="text/html",
            headers={
                "Content-Disposition": f"attachment; filename=pii_report_{case_id}.html"
            }
        )
    except Exception as e:
        logger.error(f"Failed to download report for {case_id}: {e}")
        raise HTTPException(status_code=500, detail="Error retrieving masking report.")



@router.post("/cases/{case_id}/reset")
async def reset_case(case_id: str):
    """
    Resets a case by clearing its classification and safety results.
    """
    cosmos = _get_cosmos()
    await cosmos.reset_case(case_id)
    return {"message": f"Successfully reset case {case_id}"}
